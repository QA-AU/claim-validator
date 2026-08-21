"""The per-job Excel report — Claims, Gaps, Quality sheets — built from the
same `ValidationResult` the JSON API response comes from, so the two formats
can never drift apart from each other.
"""

from typing import Any, Dict

from phases.report_style import (
    VERDICT_FILL,
    VERDICT_HUMAN_CHECK,
    VERDICT_INCORRECT,
    VERDICT_QUALITY,
    style_sheet,
    write_row,
)

from claimvalidator.pipeline import ValidationResult


def claim_verdict(result_dict: Dict[str, Any]) -> str:
    """Quality / Human check / Incorrect, for one claim row.

    Mirrors `_requirement_verdict` in the source repo's run_report_excel.py,
    adapted to this system's verdicts (there is no shape-violation-implies-
    incorrect distinction to inherit here since a shape violation on a bare
    claim already means "nothing to judge" — folded into `no_evidence`-style
    handling rather than a separate branch).
    """
    if not result_dict["shape"]["ok"]:
        return VERDICT_INCORRECT
    verdict = result_dict["verdict"]
    if verdict == "contradicts":
        return VERDICT_INCORRECT
    if verdict == "no_evidence":
        return VERDICT_INCORRECT
    if verdict == "entails":
        return VERDICT_QUALITY
    return VERDICT_HUMAN_CHECK  # mentions_only, unjudged


def build_excel_report(result: ValidationResult, path: str) -> None:
    from pathlib import Path

    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    _claims_sheet(wb, result)
    _shape_checks_sheet(wb, result)
    _gaps_sheet(wb, result)
    _quality_sheet(wb, result)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _claims_sheet(wb, result: ValidationResult) -> None:
    from openpyxl.styles import Font, PatternFill

    sheet = wb.create_sheet("Claims")
    sheet.append(["Claim ID", "Text", "Shape", "Verdict", "Quality classification",
                  "Agreement", "Escalation", "Cited chunks", "Reason"])

    r = 2
    for claim in result.per_claim:
        d = claim.to_dict()
        verdict = claim_verdict(d)
        if claim.escalated:
            escalation_note = (
                f"{claim.escalation_model}: {claim.escalated_from} → {claim.verdict}"
                if claim.escalated_from and claim.escalated_from != claim.verdict
                else f"{claim.escalation_model}: confirmed {claim.verdict}"
            )
        else:
            escalation_note = "—"
        write_row(sheet, r, [
            claim.id,
            claim.text,
            "OK" if claim.shape_ok else f"VIOLATION: {claim.shape_reason}",
            claim.verdict,
            verdict,
            claim.agreement or "-",
            escalation_note,
            ", ".join(str(c) for c in claim.cited_chunks) or "-",
            claim.reason,
        ], wrap_all_from=2)
        for col in (4, 5):
            vcell = sheet.cell(row=r, column=col)
            vcell.font = Font(name="Arial", size=10, bold=True)
            vcell.fill = PatternFill("solid", fgColor=VERDICT_FILL[verdict])
        if claim.escalated:
            ecell = sheet.cell(row=r, column=7)
            ecell.font = Font(name="Arial", size=10, bold=True, italic=True)
        r += 1

    style_sheet(sheet, 1, [12, 42, 24, 16, 20, 12, 24, 14, 42])


def _shape_pass_reason(claim) -> str:
    """Why a claim passed, not just that it did — the same courtesy a
    violation already gets. Stated from what's actually observable (the
    claim wasn't flagged, and how long its text is) rather than restating
    the active rule set verbatim, since a shape_rule_overrides caller could
    have changed what was actually required for this specific run."""
    length = len(claim.text or "")
    return f"Has a title and {length} character(s) of checkable text; not flagged by the active shape rules."


def _shape_checks_sheet(wb, result: ValidationResult) -> None:
    """The free, deterministic check broken out into its own tab, separate
    from the judge's verdict. Runs independently and unconditionally on
    every claim — a shape violation does not stop the judge from also
    running on the same claim, so a row here can show "Violation" while
    the Claims tab still shows a real judge verdict for the same id. The
    two checks answer different questions (can this be acted on at all,
    versus is it true) and are reported separately for that reason."""
    sheet = wb.create_sheet("Shape checks")
    sheet.append(["Claim ID", "Text", "Result", "Reason"])

    r = 2
    for claim in result.per_claim:
        write_row(sheet, r, [
            claim.id,
            claim.text,
            "Pass" if claim.shape_ok else "Violation",
            claim.shape_reason if not claim.shape_ok else _shape_pass_reason(claim),
        ], wrap_all_from=2)
        if not claim.shape_ok:
            from openpyxl.styles import Font, PatternFill
            cell = sheet.cell(row=r, column=3)
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = PatternFill("solid", fgColor=VERDICT_FILL[VERDICT_INCORRECT])
        r += 1

    style_sheet(sheet, 1, [12, 56, 14, 50])


def _gaps_sheet(wb, result: ValidationResult) -> None:
    sheet = wb.create_sheet("Gaps")
    sheet.append(["Concept", "Census range", "Probable instances",
                  "Addressed by claims", "Never addressed"])

    if not result.gap_report or not result.gap_report.ran:
        reason = result.gap_report.skipped_reason if result.gap_report else "not run"
        write_row(sheet, 2, ["(gap report not run)", "", "", "", reason], wrap_all_from=5)
        style_sheet(sheet, 1, [20, 14, 40, 18, 50])
        return

    r = 2
    for concept, gap in result.gap_report.per_concept.items():
        never_addressed_detail = "\n".join(
            f"{name} — {gap.never_addressed_reasons.get(name, 'reason not recorded')}"
            for name in gap.never_addressed
        ) or "(fully addressed)"
        write_row(sheet, r, [
            concept,
            f"{gap.spread_low}–{gap.spread_high}",
            ", ".join(gap.probable) or "-",
            gap.addressed_count,
            never_addressed_detail,
        ], wrap_all_from=3)
        r += 1

    style_sheet(sheet, 1, [20, 14, 34, 18, 60])


# Metric key -> (what it measures, how to read it, the verdict classification
# a *good* value at this metric represents — "—" where no single direction is
# right or wrong, e.g. a count that simply needs to match another count).
# Mirrors the explained-metric pattern the source project's own Phase 2
# Quality tab uses, adapted to this system's actual metrics — there is no
# average_score/critical_coverage here, since there is no generation to
# score; entailed/judged/contradicted/undecided/escalated/runs carry over
# because the judge itself is reused unmodified.
_METRIC_EXPLANATIONS = {
    "claims_submitted": (
        "Rows in the input file — the claims being checked.",
        "A count, not a score. Compare against `judged` below: if they differ, "
        "some claims were never evaluated.",
        "Informational",
    ),
    "shape_checked": (
        "Claims tested against the \"has enough checkable content\" rule.",
        "Should equal `claims_submitted` — the shape check runs unconditionally "
        "on every claim, never gated on anything else.",
        "Informational",
    ),
    "shape_violations": (
        "Claims with no usable text to judge — too short, or missing whatever "
        "the active rule requires (see the Shape checks tab for which claim, "
        "and the exact reason).",
        "0 is the expected value. Each one is a named, concrete defect, not a "
        "judgment call.",
        VERDICT_QUALITY,
    ),
    "retrieval_found_nothing": (
        "Claims for which no supporting passage was found anywhere in the "
        "document.",
        "Not itself a defect in the claim — it's what makes `no_evidence` a "
        "real, checked answer instead of a false `entails`. Compare against "
        "`no_evidence` below; the two should roughly track together.",
        "Informational",
    ),
    "judged": (
        "Claims actually checked against a cited or retrieved passage.",
        "0 would mean the judge never ran at all — not the same thing as "
        "every claim passing.",
        "Informational",
    ),
    "entailed": (
        "Claims the cited or retrieved passages actually support.",
        "The number this whole report exists to produce. Compare against "
        "`claims_submitted` for a rough correctness rate — but read the "
        "actual verdicts on the Claims tab before trusting the ratio alone; "
        "a claim can be entailed on a technicality that misses the point.",
        VERDICT_QUALITY,
    ),
    "mentions_only": (
        "Passages that are genuinely on-topic but silent on the specific "
        "claim made.",
        "Not automatically a defect — often means the claim is more specific "
        "than the document, or the judge correctly declined to guess rather "
        "than force an entails/contradicts call it couldn't support.",
        VERDICT_HUMAN_CHECK,
    ),
    "contradicted": (
        "Claims that conflict with the passage they cite or were matched to.",
        "The single most actionable finding this report produces — each one "
        "is a claim asserting something the document states the opposite of.",
        VERDICT_INCORRECT,
    ),
    "no_evidence": (
        "Claims with no supporting passage anywhere in the document, after "
        "retrieval genuinely looked.",
        "A plausible-sounding claim the document simply never makes. Worth "
        "checking whether the claim belongs against this document at all.",
        VERDICT_INCORRECT,
    ),
    "undecided": (
        "Verdicts the judge's own repeated runs could not settle by "
        "majority.",
        "The judge disagreeing with itself is informative on its own — an "
        "undecided claim is exactly the kind of case worth a human read, or "
        "escalation to a stronger model.",
        VERDICT_HUMAN_CHECK,
    ),
    "escalated": (
        "Doubtful verdicts (undecided, or a split contradiction) a stronger "
        "model actually answered — a completed second opinion, not merely an "
        "attempted one.",
        "Rare by design — only the cases above qualify. Read next to "
        "`overturned`: escalated with nothing overturned means the stronger "
        "model confirmed the first pass rather than corrected it. Zero here "
        "does NOT necessarily mean nothing needed escalation — check "
        "`escalation_failed_batches` below before assuming that.",
        "Informational",
    ),
    "escalation_failed_batches": (
        "Escalation calls that were attempted and failed outright — no "
        "credentials configured for the stronger tier, a network error, etc.",
        "0 alongside `escalated` = 0 means escalation was never triggered at "
        "all. Any other value alongside `escalated` = 0 means it WAS "
        "triggered and every attempt failed — the original verdicts stood "
        "by default (best-effort, never blocking), but the second opinion "
        "this run wanted was never actually obtained.",
        "Informational",
    ),
    "overturned": (
        "Verdicts the stronger model actually changed from what the first "
        "model reported.",
        "Each one is a finding the cheaper model would have reported "
        "differently — the concrete evidence that escalation did something, "
        "not just that it ran.",
        "Informational",
    ),
    "runs": (
        "How many times the judge repeated itself on each claim before "
        "taking a majority verdict.",
        "Not a quality signal by itself — it exists because a single pass is "
        "unstable. See the \"agreement\" row below for how this connects to "
        "the per-claim Agreement column on the Claims tab.",
        "Informational",
    ),
    "concepts_covered": (
        "Document concept types touched by at least one claim's citation, "
        "per the gap report.",
        "Read against `concepts_total` below. A concept type with zero "
        "coverage may mean the claims genuinely never discuss it — or that "
        "they discuss it without sharing a chunk with where the census "
        "verified a specific named instance. See the Gaps tab.",
        "Informational",
    ),
    "concepts_total": (
        "Concept types the document's own extraction discovered.",
        "The denominator for `concepts_covered` above.",
        "Informational",
    ),
    "llm_calls": (
        "Total model calls made across the whole run — ontology build (if "
        "not cached), retrieval, judging, and census.",
        "Not a cost figure by itself; see `total_tokens` and `cost_cents` "
        "below for that. Useful mainly for spotting an unusually chatty run.",
        "Informational",
    ),
    "input_tokens": (
        "Tokens sent to the model across every call this run made.",
        "Exact — reported by the provider, not estimated. Usually the "
        "larger share of `total_tokens`, since every call resends the "
        "cited passages and prompt context, not just the claim text.",
        "Informational",
    ),
    "output_tokens": (
        "Tokens the model generated across every call this run made.",
        "Exact, from the provider. Census and judge calls that ask for "
        "reasoning text cost more here than a plain verdict-only call would.",
        "Informational",
    ),
    "total_tokens": (
        "input_tokens + output_tokens for the whole run.",
        "The number that actually drives cost. Compare across runs on the "
        "same document to see what caching (ontology reuse) saved.",
        "Informational",
    ),
    "cost_cents": (
        "Estimated spend for this run, in cents, at the configured price.",
        "\"not available\" is the honest default — this project does not "
        "invent a price. Set CLAIMVAL_PRICE_INPUT_CENTS and "
        "CLAIMVAL_PRICE_OUTPUT_CENTS (cents per million tokens) to see a "
        "real estimate instead of that string.",
        "Informational",
    ),
}

# Not a real key in result.quality — agreement is inherently per-claim, shown
# as e.g. "2/3" on the Claims tab. Explained here anyway because "what does
# Agreement mean" has no other home in this report.
_AGREEMENT_ROW = (
    "agreement (per claim, see Claims tab)",
    "e.g. 2/3",
    "How many of the judge's repeated runs on ONE claim (see `runs` above) "
    "agreed on the verdict actually reported.",
    "3/3 is unanimous — the strongest confidence this system produces. 2/3 "
    "means the judge disagreed with itself once; still a majority verdict, "
    "but worth a closer read, especially on claims that turn on a fine "
    "distinction rather than a clear-cut fact. A verdict no majority could "
    "settle is reported as `undecided` above, not guessed at.",
    "Informational",
)


def _quality_sheet(wb, result: ValidationResult) -> None:
    from openpyxl.styles import Font, PatternFill

    sheet = wb.create_sheet("Quality")
    sheet.append(["Metric", "Value", "What it measures", "How to read it", "Verdict"])

    r = 2
    for key, value in result.quality.items():
        what, how, verdict = _METRIC_EXPLANATIONS.get(
            key, ("(no description on file for this metric — added since this "
                  "sheet's explanations were last written)",
                  "Treat as informational until this row is documented above.",
                  "Informational"))
        write_row(sheet, r, [key, value, what, how, verdict], wrap_all_from=3)
        if verdict in VERDICT_FILL:
            vcell = sheet.cell(row=r, column=5)
            vcell.font = Font(name="Arial", size=10, bold=True)
            vcell.fill = PatternFill("solid", fgColor=VERDICT_FILL[verdict])
        r += 1

    name, value, what, how, verdict = _AGREEMENT_ROW
    write_row(sheet, r, [name, value, what, how, verdict], wrap_all_from=3)

    style_sheet(sheet, 1, [22, 12, 34, 42, 12])
