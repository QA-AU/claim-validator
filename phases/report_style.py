"""Low-level Excel styling, extracted from the source repo's
`phases/run_report_excel.py` — the constants and two writer functions every
sheet in that module builds on, without the sheet-builder functions
themselves (`_requirements_sheet` etc.), which are shaped around that
pipeline's own artifacts and don't fit a bare claim list.

Kept so this repo's reports *look and read* like the project it was built
from — same font, same header styling, same Quality/Human check/Incorrect
verdict vocabulary and colours — without inheriting code that doesn't apply
here.
"""

from typing import Any, List

FONT = "Arial"

HEAD_FILL = "1F4E79"
WARN_FILL = "FDF3F2"
OK_FILL = "F2F7F2"
WARN_INK = "A0302A"
MUTED = "5F6B76"

# Same three-way classification the source repo uses on every item sheet —
# reducing whichever checks ran on a row (judge verdict, shape violation,
# retrieval outcome) to the one question a reviewer scanning many rows asks
# first. Each signal that produced it stays in its own column beside it.
#
#   Quality       every check that ran on this item came back clean
#   Human check   nothing confirmed wrong, but something is unverified —
#                 unjudged, undecided, no citation found, or genuinely
#                 on-topic-but-silent (mentions_only)
#   Incorrect     a check found a concrete defect: a contradiction, a shape
#                 violation, a claim with no supporting passage anywhere
VERDICT_QUALITY = "Quality"
VERDICT_HUMAN_CHECK = "Human check"
VERDICT_INCORRECT = "Incorrect"

VERDICT_FILL = {
    VERDICT_QUALITY: "D9EAD3",
    VERDICT_HUMAN_CHECK: "FCE5CD",
    VERDICT_INCORRECT: "F4CCCC",
}


def style_sheet(ws, header_row: int, widths: List[int]) -> None:
    """Header font/fill, column widths, frozen header row."""
    from openpyxl.styles import Alignment, Font, PatternFill

    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    for cell in ws[header_row]:
        if cell.value is None:
            continue
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=HEAD_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


# Formula-injection guard (CWE-1236): a string cell starting with any of
# these gets evaluated as a formula by Excel/LibreOffice/Sheets once the
# file is opened. Every row through here can carry untrusted API input —
# claim.text and claim.id are caller-supplied — so this isn't optional.
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _defuse(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(_FORMULA_TRIGGER_CHARS):
        return "'" + value
    return value


def write_row(ws, row: int, values: List[Any], wrap_last: bool = False,
              wrap_all_from: "int | None" = None) -> None:
    """`wrap_all_from` wraps every column from that index on, for prose columns."""
    from openpyxl.styles import Alignment, Font

    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=_defuse(value))
        cell.font = Font(name=FONT, size=10)
        wrap = (wrap_last and col == len(values)) or (
            wrap_all_from is not None and col >= wrap_all_from
        )
        cell.alignment = Alignment(vertical="top", wrap_text=wrap)
