"""claimvalidator/report_excel.py's pure helper functions — the ones that
decide what a report cell actually says, independent of openpyxl. Found
worth testing directly after a report reader couldn't tell which claim IDs
a Quality tab metric like "contradicted: 2" was actually about.
"""

from claimvalidator.pipeline import ClaimResult, ValidationResult
from claimvalidator.report_excel import _claim_ids_for, _claims_sheet


def _claim(id, **overrides):
    defaults = dict(
        id=id, text="t", shape_ok=True, shape_reason=None, verdict="entails",
        judged=True, agreement="3/3", cited_chunks=[0], reason="r",
    )
    defaults.update(overrides)
    return ClaimResult(**defaults)


def test_claims_sheet_writes_a_source_ref_column():
    from openpyxl import Workbook

    claims = [
        _claim("C1", source_ref="README.md, para 3"),
        _claim("C2"),  # no source_ref — the common case, must not break
    ]
    result = ValidationResult(ontology_key="k", ontology_reused=False, per_claim=claims)

    wb = Workbook()
    _claims_sheet(wb, result)
    sheet = wb["Claims"]

    header = [c.value for c in sheet[1]]
    assert header[-1] == "Source ref"

    assert sheet.cell(row=2, column=len(header)).value == "README.md, para 3"
    # "-" alone starts with a formula-trigger character (write_row's own
    # defuse guard, see phases/report_style.py) — "'-" is the correct,
    # already-defused value, not a bug in this new column.
    assert sheet.cell(row=3, column=len(header)).value == "'-"


def test_a_metric_with_no_claim_id_meaning_says_so_rather_than_blank():
    assert _claim_ids_for([_claim("C1")], "llm_calls") == "(whole-run figure, not per-claim)"


def test_a_zero_count_metric_reports_none_not_an_empty_string():
    claims = [_claim("C1", verdict="entails")]
    assert _claim_ids_for(claims, "contradicted") == "(none)"


def test_contradicted_names_only_the_contradicting_claims():
    claims = [
        _claim("C1", verdict="entails"),
        _claim("C2", verdict="contradicts"),
        _claim("C3", verdict="contradicts"),
    ]
    assert _claim_ids_for(claims, "contradicted") == "C2, C3"


def test_shape_violations_names_claims_that_failed_the_shape_check():
    claims = [_claim("C1", shape_ok=False), _claim("C2", shape_ok=True)]
    assert _claim_ids_for(claims, "shape_violations") == "C1"


def test_retrieval_found_nothing_names_claims_with_no_cited_chunks():
    claims = [_claim("C1", cited_chunks=[]), _claim("C2", cited_chunks=[3])]
    assert _claim_ids_for(claims, "retrieval_found_nothing") == "C1"


def test_undecided_names_judged_claims_the_judge_could_not_settle():
    claims = [
        _claim("C1", judged=True, decided=True),
        _claim("C2", judged=True, decided=False),
        _claim("C3", judged=False, decided=False),  # unjudged, not "undecided"
    ]
    assert _claim_ids_for(claims, "undecided") == "C2"


def test_overturned_names_only_escalated_claims_whose_verdict_actually_changed():
    claims = [
        _claim("C1", escalated=True, escalated_from="entails", verdict="entails"),  # confirmed, not overturned
        _claim("C2", escalated=True, escalated_from="mentions_only", verdict="contradicts"),
        _claim("C3", escalated=False),
    ]
    assert _claim_ids_for(claims, "overturned") == "C2"
