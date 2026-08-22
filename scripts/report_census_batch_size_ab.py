"""Excel report for the census batch-size A/B — not a run_validation() job
(no claims were involved; this measured census_repeated() directly), so it
never went through report_excel.py's build_excel_report(). Reconstructed
here from the actual recorded output of the completed run, rather than
left as a log file nobody but this repo's history can reread.

    python scripts/report_census_batch_size_ab.py

The counts below are the real recorded result of one specific run: RFC
6749, Anthropic claude-haiku-4-5, 2026-08-22, 3 runs per condition,
batch_size=10 vs batch_size=3. See docs/Independent-Claim-Validator.docx
for the full narrative — three failed attempts (gpt-oss:120b-cloud's quota,
two local models hitting a 600s timeout cliff, and a previously-latent
AnthropicClient bug the switch to Anthropic surfaced) preceded this clean
run. Re-running the same comparison on a different model or document would
need new counts substituted below, not a code change — this script reports
one recorded measurement, it does not re-run the census itself.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from phases.report_style import (
    VERDICT_FILL,
    VERDICT_HUMAN_CHECK,
    VERDICT_INCORRECT,
    VERDICT_QUALITY,
    style_sheet,
    write_row,
)

DOCUMENT = "RFC 6749 (rfc6749-8b06)"
MODEL = "claude-haiku-4-5 (Anthropic)"
RUN_DATE_UTC = "2026-08-22"

# Real recorded counts, one census_repeated() call per condition, 3 runs
# each. Temperature was requested at 0.01 (census.py's pin) but silently
# dropped by the installed anthropic SDK, which no longer accepts it — see
# AnthropicClient's fix, committed the same day — so this ran at the
# server's own default sampling, not a pinned low temperature.
BATCH_10 = {
    "role": [15, 16, 16],
    "grant_type": [20, 19, 17],
    "error_code": [16, 19, 17],
    "token": [8, 11, 14],
    "parameter": [50, 43, 46],
    "http_response": [22, 18, 21],
}
BATCH_3 = {
    "role": [25, 31, 31],
    "grant_type": [25, 28, 24],
    "error_code": [25, 31, 24],
    "token": [24, 22, 24],
    "parameter": [85, 78, 70],
    "http_response": [39, 39, 39],
}

USAGE = {"calls": 240, "input_tokens": 363210, "output_tokens": 86334}
# claude-haiku-4-5: $1.00 / $5.00 per MTok, input / output.
COST_DOLLARS = (USAGE["input_tokens"] * 1.00 + USAGE["output_tokens"] * 5.00) / 1_000_000


def _spread(counts):
    return max(counts) - min(counts)


def _spread_verdict(delta: int) -> str:
    if delta == 0:
        return VERDICT_QUALITY
    if delta <= 3:
        return VERDICT_HUMAN_CHECK
    return VERDICT_INCORRECT


def _shade(cell, verdict: str) -> None:
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.fill = PatternFill("solid", fgColor=VERDICT_FILL[verdict])


def _batch_size_sheet(wb) -> tuple:
    sheet = wb.create_sheet("Batch size AB")
    sheet.append(["Concept", "batch_size=10 counts", "b10 spread", "batch_size=3 counts",
                  "b3 spread", "spread delta (b3 - b10)"])

    r = 2
    b10_sum = b3_sum = 0
    for name in BATCH_10:
        c10, c3 = BATCH_10[name], BATCH_3[name]
        s10, s3 = _spread(c10), _spread(c3)
        b10_sum += s10
        b3_sum += s3
        delta = s3 - s10
        write_row(sheet, r, [
            name,
            ", ".join(str(x) for x in c10), f"{min(c10)}-{max(c10)} ({s10})",
            ", ".join(str(x) for x in c3), f"{min(c3)}-{max(c3)} ({s3})",
            delta,
        ])
        _shade(sheet.cell(row=r, column=6), _spread_verdict(abs(delta)))
        r += 1

    write_row(sheet, r, ["TOTAL", "", f"sum={b10_sum}", "", f"sum={b3_sum}", b3_sum - b10_sum])
    for col in range(1, 7):
        sheet.cell(row=r, column=col).font = Font(name="Arial", size=10, bold=True)

    style_sheet(sheet, 1, [16, 22, 16, 22, 16, 20])
    return b10_sum, b3_sum


def _summary_sheet(wb, b10_sum: int, b3_sum: int) -> None:
    sheet = wb.create_sheet("Summary")
    sheet.append(["Field", "Value", "Note"])

    zero_spread_10 = sum(1 for c in BATCH_10.values() if _spread(c) == 0)
    zero_spread_3 = sum(1 for c in BATCH_3.values() if _spread(c) == 0)

    rows = [
        ("Document", DOCUMENT, "182 chunks, 6 concept types"),
        ("Model", MODEL, "not comparable to the gpt-oss:120b-cloud temperature numbers "
         "elsewhere in this project — different model, run separately on purpose"),
        ("Runs per condition", "3", "census_repeated's standard majority-consensus repeat count"),
        ("Temperature requested", "0.01", "silently dropped by the installed anthropic SDK — "
         "Messages.create() no longer accepts it; this run used the server's own default "
         "sampling, not a pinned low temperature"),
        ("Sum of (high-low), batch_size=10", b10_sum, "baseline"),
        ("Sum of (high-low), batch_size=3", b3_sum,
         f"{'+' if b3_sum > b10_sum else ''}{b3_sum - b10_sum} vs baseline — spread got WORSE "
         "at the smaller batch size, the opposite of the hypothesis this A/B tested"),
        ("Concepts with zero spread, b10", zero_spread_10, "out of 6"),
        ("Concepts with zero spread, b3", zero_spread_3, "out of 6"),
        ("LLM calls", USAGE["calls"],
         "both conditions combined: 19 batches x3 runs (b10) + 61 batches x3 runs (b3)"),
        ("Input tokens", USAGE["input_tokens"], ""),
        ("Output tokens", USAGE["output_tokens"], "higher than a pre-run estimate by 7-10x — "
         "this model found substantially more instances than assumed, and listing more "
         "instances costs more output tokens"),
        ("Estimated cost", f"${COST_DOLLARS:.2f}",
         "at claude-haiku-4-5 pricing ($1.00/$5.00 per MTok, input/output)"),
        ("Run date", RUN_DATE_UTC, "UTC"),
    ]
    for i, (field, value, note) in enumerate(rows, start=2):
        write_row(sheet, i, [field, value, note], wrap_all_from=3)

    style_sheet(sheet, 1, [30, 24, 60])


def _interpretation_sheet(wb) -> None:
    sheet = wb.create_sheet("What this means")
    sheet.append(["Finding", "Detail", "Verdict"])

    findings = [
        ("Smaller batch increased total spread, not decreased it",
         "24 -> 34 (b10 -> b3), the opposite of the recall-dilution hypothesis this A/B was "
         "built to test.", VERDICT_INCORRECT),
        ("Absolute counts rose substantially at the smaller batch size",
         "e.g. parameter: 43-50 -> 70-85; http_response: 18-22 -> 39 every run. The smaller "
         "batch is finding real instances the larger one was missing.", VERDICT_QUALITY),
        ("The recall gain is not landing consistently run to run",
         "More real instances found, but with more disagreement about exactly how many, on "
         "most concepts — a wider absolute spread partly reflecting a more complete picture "
         "measured less reliably.", VERDICT_HUMAN_CHECK),
        ("Effect is concept-dependent, not uniform",
         "http_response went from unstable (18-22) to perfectly stable (39 every run) at the "
         "smaller batch size — the same concept-by-concept inconsistency every instability "
         "finding in this project has shown.", VERDICT_HUMAN_CHECK),
        ("Not yet known whether this generalises",
         "One model (claude-haiku-4-5), one document (RFC 6749). Whether gpt-oss:120b-cloud "
         "shows the same inversion once its quota resets, or a different document shows the "
         "same pattern, is untested.", VERDICT_HUMAN_CHECK),
    ]
    for i, (finding, detail, verdict) in enumerate(findings, start=2):
        write_row(sheet, i, [finding, detail, verdict], wrap_all_from=2)
        _shade(sheet.cell(row=i, column=3), verdict)

    style_sheet(sheet, 1, [34, 60, 14])


def main():
    wb = Workbook()
    wb.remove(wb.active)

    b10_sum, b3_sum = _batch_size_sheet(wb)
    _summary_sheet(wb, b10_sum, b3_sum)
    _interpretation_sheet(wb)

    path = "reports/census_batch_size_ab_rfc6749.xlsx"
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"written: {path}")


if __name__ == "__main__":
    main()
